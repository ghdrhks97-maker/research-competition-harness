"""Survey analysis engine (survey-analysis-skill / `rch import-survey`).

Reads a pre/post survey table (CSV, TSV, or XLSX) and produces an
anonymized analysis: descriptive stats, pre/post change, Cohen's d, a
two-sided t-test p-value, a free-response summary, and honest
small-sample caveats. Personally identifying columns are dropped and
reported, never analyzed. Numeric findings become `derived` claims;
free-text quotes stay `placeholder` until a human confirms consent.
"""

from __future__ import annotations

import csv
import json
import math
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from rch import stats

# Header substrings that mark a column as personally identifying. Such
# columns are excluded from analysis and listed in the privacy report.
PII_MARKERS = (
    "이름",
    "성명",
    "학번",
    "번호",
    "연락처",
    "전화",
    "휴대",
    "email",
    "e-mail",
    "메일",
    "주소",
    "name",
    "phone",
    "student id",
    "studentid",
)

PRE_MARKERS = ("사전", "pre", "이전", "before", "전")
POST_MARKERS = ("사후", "post", "이후", "after", "후")

STOPWORDS = {
    "그리고", "그러나", "하지만", "그래서", "그런데", "정말", "진짜", "너무",
    "매우", "조금", "그냥", "좀", "것", "수", "등", "및", "the", "and", "a",
    "to", "of", "is", "it", "i", "was", "were", "이", "그", "저", "때",
}

SMALL_SAMPLE_THRESHOLD = 30


@dataclass
class ItemResult:
    item: str
    kind: str  # "paired" | "descriptive"
    n: int
    pre_mean: float | None = None
    post_mean: float | None = None
    mean_diff: float | None = None
    sd_diff: float | None = None
    cohens_d: float | None = None
    effect_label: str | None = None
    t: float | None = None
    df: int | None = None
    p_value: float | None = None
    mean: float | None = None
    sd: float | None = None
    note: str = ""


@dataclass
class FreeResponse:
    column: str
    response_count: int
    top_keywords: list[list[Any]]  # [keyword, count]
    sample_quotes: list[str]


@dataclass
class SurveyAnalysis:
    source: str
    respondents: int
    dropped_pii_columns: list[str] = field(default_factory=list)
    paired_items: list[ItemResult] = field(default_factory=list)
    descriptive_items: list[ItemResult] = field(default_factory=list)
    free_responses: list[FreeResponse] = field(default_factory=list)
    limitations: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        return data


def _read_table(path: Path) -> tuple[list[str], list[list[str]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    delimiter = "\t" if suffix in {".tsv", ".tab"} else ","
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        raise ValueError(f"empty survey file: {path}")
    return rows[0], rows[1:]


def _read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised only without dep
        raise ValueError(
            "XLSX 분석에는 openpyxl이 필요합니다. `pip install openpyxl` 후 다시 실행하거나 CSV로 저장하세요."
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[list[str]] = []
    for raw in sheet.iter_rows(values_only=True):
        cells = ["" if value is None else str(value) for value in raw]
        if any(cell.strip() for cell in cells):
            rows.append(cells)
    if not rows:
        raise ValueError(f"empty survey file: {path}")
    return rows[0], rows[1:]


def _is_pii(header: str) -> bool:
    low = header.strip().lower()
    return any(marker in low for marker in PII_MARKERS)


def _to_float(value: str) -> float | None:
    value = value.strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _column_values(rows: list[list[str]], index: int) -> list[str]:
    return [row[index] if index < len(row) else "" for row in rows]


def _numeric_column(rows: list[list[str]], index: int) -> tuple[list[float], int]:
    """Return numeric values and count of non-empty-but-non-numeric cells."""
    numbers: list[float] = []
    non_numeric = 0
    for cell in _column_values(rows, index):
        stripped = cell.strip()
        if not stripped:
            continue
        parsed = _to_float(stripped)
        if parsed is None:
            non_numeric += 1
        else:
            numbers.append(parsed)
    return numbers, non_numeric


def _base_name(header: str) -> str:
    base = header.strip()
    for marker in PRE_MARKERS + POST_MARKERS:
        base = re.sub(rf"[\s_\-]*{re.escape(marker)}[\s_\-]*", "", base, flags=re.IGNORECASE)
    return base.strip(" _-").strip() or header.strip()


def _phase(header: str) -> str | None:
    low = header.strip().lower()
    if any(marker in low for marker in PRE_MARKERS):
        return "pre"
    if any(marker in low for marker in POST_MARKERS):
        return "post"
    return None


def _round(value: float | None, digits: int = 3) -> float | None:
    if value is None or not isinstance(value, float) or not math.isfinite(value):
        return None
    return round(value, digits)


def _summarize_free_text(column: str, values: list[str]) -> FreeResponse:
    responses = [value.strip() for value in values if value.strip()]
    counter: dict[str, int] = {}
    for response in responses:
        for token in re.findall(r"[0-9A-Za-z가-힣]+", response.lower()):
            if len(token) < 2 or token in STOPWORDS:
                continue
            counter[token] = counter.get(token, 0) + 1
    top = sorted(counter.items(), key=lambda item: (-item[1], item[0]))[:10]
    # Representative quotes: shortest distinct responses, kept verbatim but
    # flagged for privacy review downstream.
    seen: set[str] = set()
    quotes: list[str] = []
    for response in sorted(responses, key=len):
        if response in seen:
            continue
        seen.add(response)
        quotes.append(response)
        if len(quotes) >= 3:
            break
    return FreeResponse(
        column=column,
        response_count=len(responses),
        top_keywords=[[keyword, count] for keyword, count in top],
        sample_quotes=quotes,
    )


def analyze_table(headers: list[str], rows: list[list[str]], source: str) -> SurveyAnalysis:
    analysis = SurveyAnalysis(source=source, respondents=len(rows))

    keep_indexes: list[int] = []
    for index, header in enumerate(headers):
        if _is_pii(header):
            analysis.dropped_pii_columns.append(header.strip())
        else:
            keep_indexes.append(index)

    # Group numeric columns by base item name to detect pre/post pairs.
    numeric_indexes: list[int] = []
    text_indexes: list[int] = []
    for index in keep_indexes:
        numbers, non_numeric = _numeric_column(rows, index)
        if numbers and non_numeric == 0:
            numeric_indexes.append(index)
        elif any(cell.strip() for cell in _column_values(rows, index)):
            text_indexes.append(index)

    groups: dict[str, dict[str, int]] = {}
    solo_numeric: list[int] = []
    for index in numeric_indexes:
        phase = _phase(headers[index])
        if phase is None:
            solo_numeric.append(index)
            continue
        base = _base_name(headers[index])
        groups.setdefault(base, {})[phase] = index

    for base, phases in groups.items():
        if "pre" in phases and "post" in phases:
            analysis.paired_items.append(_paired_item(base, headers, rows, phases))
        else:
            for index in phases.values():
                solo_numeric.append(index)

    for index in solo_numeric:
        numbers, _ = _numeric_column(rows, index)
        analysis.descriptive_items.append(
            ItemResult(
                item=headers[index].strip(),
                kind="descriptive",
                n=len(numbers),
                mean=_round(stats.mean(numbers)),
                sd=_round(stats.stdev(numbers)),
            )
        )

    for index in text_indexes:
        analysis.free_responses.append(
            _summarize_free_text(headers[index].strip(), _column_values(rows, index))
        )

    analysis.limitations = _build_limitations(analysis)
    return analysis


def _paired_item(
    base: str, headers: list[str], rows: list[list[str]], phases: dict[str, int]
) -> ItemResult:
    pre_index = phases["pre"]
    post_index = phases["post"]
    pre: list[float] = []
    post: list[float] = []
    for row in rows:
        pre_value = _to_float(row[pre_index]) if pre_index < len(row) else None
        post_value = _to_float(row[post_index]) if post_index < len(row) else None
        if pre_value is None or post_value is None:
            continue  # listwise deletion for paired analysis
        pre.append(pre_value)
        post.append(post_value)
    result = stats.paired_test(pre, post)
    note = ""
    if result.n < SMALL_SAMPLE_THRESHOLD:
        note = f"표본 {result.n}명으로 소표본. p값·효과크기는 참고용."
    return ItemResult(
        item=base,
        kind="paired",
        n=result.n,
        pre_mean=_round(result.pre_mean),
        post_mean=_round(result.post_mean),
        mean_diff=_round(result.mean_diff),
        sd_diff=_round(result.sd_diff),
        cohens_d=_round(result.cohens_d),
        effect_label=stats.cohens_d_interpretation(result.cohens_d),
        t=_round(result.t),
        df=result.df,
        p_value=_round(result.p_value, 4),
        note=note,
    )


def _build_limitations(analysis: SurveyAnalysis) -> list[str]:
    limitations: list[str] = []
    if analysis.respondents < SMALL_SAMPLE_THRESHOLD:
        limitations.append(
            f"전체 응답자 {analysis.respondents}명은 소표본이므로 통계적 유의성보다 변화 경향으로 해석한다."
        )
    if analysis.dropped_pii_columns:
        limitations.append(
            "개인정보로 판단된 열은 분석에서 제외했고 원자료는 부록·commit에 포함하지 않는다."
        )
    significant = [
        item
        for item in analysis.paired_items
        if item.p_value is not None and item.p_value >= 0.05
    ]
    if significant:
        limitations.append(
            "일부 문항은 통계적으로 유의하지 않다. 유의하지 않은 결과도 그대로 보고한다."
        )
    if analysis.free_responses:
        limitations.append(
            "자유응답 인용은 사용 전 학생 동의와 익명화를 사람 검토로 확정해야 한다."
        )
    if not limitations:
        limitations.append("분석 한계 없음으로 단정하지 말고 표본과 맥락을 함께 보고한다.")
    return limitations


def render_summary_markdown(analysis: SurveyAnalysis) -> str:
    lines = ["# 설문 분석 요약", "", f"- 출처: `{analysis.source}`", f"- 응답자: {analysis.respondents}명"]
    if analysis.dropped_pii_columns:
        lines.append(f"- 개인정보로 제외한 열: {', '.join(analysis.dropped_pii_columns)}")
    lines.append("")

    if analysis.paired_items:
        lines += ["## 사전·사후 변화", "", "| 문항 | n | 사전 | 사후 | 변화 | Cohen's d | 효과 | p |", "| --- | --- | --- | --- | --- | --- | --- | --- |"]
        for item in analysis.paired_items:
            lines.append(
                "| {item} | {n} | {pre} | {post} | {diff} | {d} | {effect} | {p} |".format(
                    item=item.item,
                    n=item.n,
                    pre=_fmt(item.pre_mean),
                    post=_fmt(item.post_mean),
                    diff=_fmt(item.mean_diff),
                    d=_fmt(item.cohens_d),
                    effect=item.effect_label or "-",
                    p=_fmt(item.p_value),
                )
            )
        lines.append("")

    if analysis.descriptive_items:
        lines += ["## 단일 문항 기술통계", "", "| 문항 | n | 평균 | 표준편차 |", "| --- | --- | --- | --- |"]
        for item in analysis.descriptive_items:
            lines.append(f"| {item.item} | {item.n} | {_fmt(item.mean)} | {_fmt(item.sd)} |")
        lines.append("")

    if analysis.free_responses:
        lines += ["## 자유응답", ""]
        for free in analysis.free_responses:
            keywords = ", ".join(f"{word}({count})" for word, count in free.top_keywords) or "-"
            lines.append(f"### {free.column} (응답 {free.response_count}건)")
            lines.append("")
            lines.append(f"- 상위 키워드: {keywords}")
            lines.append("- 대표 응답(동의·익명화 확인 필요):")
            for quote in free.sample_quotes:
                lines.append(f"  - \"{quote}\"")
            lines.append("")

    lines += ["## 분석 한계", ""]
    lines += [f"- {item}" for item in analysis.limitations]
    lines.append("")
    return "\n".join(lines)


def _fmt(value: float | None) -> str:
    if value is None:
        return "-"
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def build_claim_ledger(analysis: SurveyAnalysis) -> dict[str, Any]:
    claims: list[dict[str, Any]] = []
    for index, item in enumerate(analysis.paired_items, 1):
        direction = "상승" if (item.mean_diff or 0) > 0 else "하락" if (item.mean_diff or 0) < 0 else "변화 없음"
        claims.append(
            {
                "id": f"survey-paired-{index}",
                "text": f"'{item.item}' 문항 사전 {_fmt(item.pre_mean)} → 사후 {_fmt(item.post_mean)} ({direction}, d={_fmt(item.cohens_d)})",
                "status": "derived",
                "evidence": "analysis/survey-analysis.json",
                "notes": f"paired t-test, n={item.n}. {item.note}".strip(),
            }
        )
    for index, item in enumerate(analysis.descriptive_items, 1):
        claims.append(
            {
                "id": f"survey-desc-{index}",
                "text": f"'{item.item}' 평균 {_fmt(item.mean)} (n={item.n})",
                "status": "derived",
                "evidence": "analysis/survey-analysis.json",
            }
        )
    for index, free in enumerate(analysis.free_responses, 1):
        claims.append(
            {
                "id": f"survey-free-{index}",
                "text": f"'{free.column}' 자유응답 인용 후보",
                "status": "placeholder",
                "notes": "학생 동의·익명화 확인 후 real/derived로 승격.",
            }
        )
    return {"claims": claims}


def import_survey(input_path: Path, output_dir: Path) -> SurveyAnalysis:
    headers, rows = _read_table(input_path)
    analysis = analyze_table(headers, rows, source=input_path.name)
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "survey-analysis.json").write_text(
        json.dumps(analysis.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (output_dir / "survey-summary.md").write_text(
        render_summary_markdown(analysis), encoding="utf-8"
    )
    (output_dir / "claim-ledger.json").write_text(
        json.dumps(build_claim_ledger(analysis), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return analysis
